import requests
from bs4 import BeautifulSoup
import urllib.request as req
import os
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.styles import PatternFill, Alignment, Border, Side

# 기존 엑셀 파일이 있으면 삭제
if os.path.exists("./Best novel.xlsx"):
    os.remove("./Best novel.xlsx")

# 엑셀 파일 생성
book = openpyxl.Workbook()
sheet = book.active

# 웹 페이지에서 HTML 가져오기
code = requests.get("https://novel.naver.com/best/ranking")
soup = BeautifulSoup(code.text, "html.parser")

# 제목, 이미지, 저자, 점수 선택
title_elements = soup.select("span.title")
image_elements = soup.select("div.thumbnail > img")
author_elements = soup.select("span.author")
score_elements = soup.select("span.score_area")

# 이미지 다운로드 폴더 생성
folder_name = "./Best novel"
if not os.path.exists(folder_name):
    os.mkdir(folder_name)

# 엑셀 시트에 제목, 저자, 점수 기록
sheet.cell(row=1, column=1).value = "책 이미지"
sheet.cell(row=1, column=2).value = "책 제목"
sheet.cell(row=1, column=3).value = "책 저자"
sheet.cell(row=1, column=4).value = "별점"
sheet.column_dimensions["A"].width = 21
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 20
sheet.column_dimensions["D"].width = 20

# 첫 행의 높이 설정
sheet.row_dimensions[1].height = 16  

# 첫 행의 셀 색상 및 가운데 정렬 설정
header_fill = PatternFill(patternType="solid", fgColor="4FC142")
for cell in sheet[1]:
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")  # 가운데 정렬

num = 1    
for title, image, author, score in zip(title_elements, image_elements, author_elements, score_elements):
    print(f"제목: {title.text}")
    img_url = image.attrs["src"]
    
    # 이미지 다운로드 및 저장
    img_path = f"{folder_name}/{num}.jpg"
    req.urlretrieve(img_url, img_path)
    
    # 이미지 리사이즈
    img = PILImage.open(img_path)
    img_resized = img.resize((int(img.width * 0.6), int(img.height * 0.6)))
    if img_resized.mode != 'RGB':
        img_resized = img_resized.convert('RGB')
    resized_img_path = f"{folder_name}/{num}_resized.jpg"
    img_resized.save(resized_img_path)
    img.close()
    
    # 엑셀에 이미지 첨부
    sheet.add_image(Image(resized_img_path), f"A{num+1}")
    sheet.cell(row=num+1, column=2).value = title.text
    sheet.cell(row=num+1, column=3).value = author.text
    sheet.cell(row=num+1, column=4).value = score.text.strip()
    sheet.row_dimensions[num+1].height = 100
    num += 1
    print("----------------------------")

# 전체 셀에 정렬 적용
alignment_center = Alignment(horizontal="center", vertical="center")
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = alignment_center

# 테두리 설정
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 모든 셀에 테두리 적용
for row in sheet.iter_rows():
    for cell in row:
        cell.border = border

# 엑셀 파일 저장
book.save("./Best novel.xlsx")
